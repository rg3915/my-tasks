import json
import os
from collections import defaultdict
from datetime import date, datetime, timedelta

import gitlab
from decouple import config
from openpyxl import Workbook, load_workbook, styles
from rich import print

from backend.core.utils import datetime_to_string, timedelta_to_string
from backend.task.models import Issue, Label, Sprint, Task, Timesheet

FOLDER_BASE = '/home/regis/Dropbox/projetos'


def check_if_the_date_already_exists(filename, milestone_title):
    date_format = datetime.now().strftime('%Y-%m-%d')

    # Check if the file exists, and create it if it doesn't
    if not os.path.exists(filename):
        with open(filename, 'w') as file:
            file.write(f"## {milestone_title}\n")

    # Check if the date already exists in the file
    with open(filename, 'r') as file:
        file_contents = file.read()
        if date_format in file_contents:
            print("Date already exists in the file.")
        else:
            # Append the date to the file
            with open(filename, 'a') as file:
                file.write(f"\n## {date_format}\n\n")
            print(f"Date {date_format} added to the file {filename}.")


def write_on_tarefas(filename, issue, labels, is_bug):
    today = date.today().strftime('%d/%m/%y')

    _labels = ','.join(labels)

    with open(filename, 'a') as f:
        f.write('\n---\n\n')
        f.write(f'[ ] {issue.number} - {issue.title}\n')
        f.write(f'    {_labels}\n')
        f.write(f'    {today}\n\n\n')

        if issue.description:
            f.write(f'    {issue.description}\n\n')

        title = issue.title
        if is_bug:
            title = f'bugfix: {title}'

        f.write(f"    _gadd '{title}. close #{issue.number}'; # gp\n")


def write_changelog_dropbox(issue):
    customer = issue.sprint.project.customer.name
    project = issue.sprint.project.title
    filename = f'{FOLDER_BASE}/{customer}/{project}/changelog/CHANGELOG_{issue.milestone.title}.md'
    print(filename)

    check_if_the_date_already_exists(filename, issue.milestone.title)

    text = f'* {issue.title}. #{issue.number}\n'

    with open(filename, 'r') as f:
        content = f.read()
        # Escreve somente se o texto não existir.
        if text not in content:
            with open(filename, 'a') as f:
                f.write(text)


def save_issue(data):
    labels = Label.objects.filter(label__in=data['labels'])
    sprint = Sprint.objects.filter(project=data['project']).last()

    issue = Issue.objects.create(
        number=data['iid'],
        title=data['title'],
        description=data['description'],
        milestone=data['milestone'],
        sprint=sprint,
        url=data['web_url'],
    )
    for label in labels:
        issue.labels.add(label)

    filename = f'{FOLDER_BASE}/{sprint.project.customer.name}/{sprint.project.title}/tarefas.txt'

    number = data['iid']
    title = data['title']

    print(filename)
    print(number)
    print(title)

    is_bug = False
    if 'bug' in data['labels']:
        is_bug = True

    write_on_tarefas(filename, issue, data['labels'], is_bug)
    return issue


def save_task(issue):
    data = dict(
        title=issue.title,
        project=issue.sprint.project,
        issue=issue,
    )
    Task.objects.create(**data)


def create_timesheet(task, previous_hour):
    now = datetime.now()
    start_time = datetime.now()

    if previous_hour:
        # TODO: verificar se é do mesmo projeto.
        last_hour = Timesheet.objects.first()  # é o último por causa da ordenação.
        start_time = last_hour.end_time
        print('start:', datetime_to_string(start_time - timedelta(hours=3), '%H:%M'))
        print('now:', datetime_to_string(now, '%H:%M'))
    else:
        print(datetime_to_string(start_time, '%H:%M'))

    return Timesheet.objects.create(task=task, start_time=start_time)


def stop_timesheet(task):
    now = datetime.now()

    print(datetime_to_string(now, '%H:%M'))

    # Na verdade não precisa da task.
    timesheet = Timesheet.objects.filter(task=task, end_time__isnull=True).first()
    timesheet.end_time = now
    timesheet.save()


def create_gitlab_issue(args):
    '''
    Requer /etc/rg3915.cfg
    '''
    gl = gitlab.Gitlab.from_config('somewhere', ['/etc/rg3915.cfg'])

    title, body, labels, project, milestone = args.values()

    gl_project = gl.projects.get(project.gitlab_project_id)

    data_dict = {
        "title": f"{title}",
        "description": f"{body}",
        "assignee_id": config('GITLAB_ASSIGNEE_ID'),
        "labels": labels,
        "milestone_id": milestone.original_id,
    }

    response = gl_project.issues.create(data_dict)

    data = json.loads(response.to_json())

    # TESTE
    # data = {
    #     "iid": 7,
    #     "title": "Teste",
    #     "description": "Descrição teste.",
    #     "labels": ["backend", "frontend"],
    #     # "time_stats":
    #     # {
    #     #     "time_estimate": 0,
    #     #     "total_time_spent": 0,
    #     # },
    #     "web_url": "https://gitlab.com/rg3915/my-tasks/-/issues/7",
    # }

    data['project'] = project
    data['milestone'] = milestone

    return data


def get_hour_display(_time):
    hours, remainder = divmod(_time.total_seconds(), 3600)
    minutes = divmod(remainder, 60)
    time_str = ''

    if hours:
        time_str += f'{int(hours)}h '

    if minutes[0]:
        time_str += f'{int(minutes[0])}m'

    if not time_str:
        return '0'

    return time_str.strip()


def group_by_date(project):
    '''
    Agrupa os dados por dia.
    '''
    timesheet_data = Timesheet.objects.filter(task__project__title=project).values(
        'start_time__date',
        'task__issue__number',
        'start_time',
        'end_time',
        'task__issue__sprint__number',
    ).order_by('start_time')

    # Create a dictionary to store the total hours and issues for each date
    result_dict = defaultdict(lambda: {'total_hours': timedelta(), 'issues': set()})

    for timesheet in timesheet_data:
        date_only = timesheet['start_time'].date()
        total_hours = timesheet['end_time'] - timesheet['start_time']
        issues = timesheet['task__issue__number']
        result_dict[date_only]['total_hours'] += total_hours
        result_dict[date_only]['issues'].add(issues)
        result_dict[date_only]['sprint'] = timesheet['task__issue__sprint__number']

    # Convert the dictionary to the desired output format
    output = [
        {
            'date': datetime_to_string(key, '%d/%m/%y'),
            'month': key.month,
            'total_hours': timedelta_to_string(value['total_hours']),
            'total_hours_display': str(get_hour_display(value['total_hours'])),
            'issues': ', '.join(map(str, value['issues'])),
            'sprint': value['sprint'],
        }
        for key, value in result_dict.items()
    ]

    return output


def write_total_hours_on_timesheet_file(customer, project, total_hours):
    timesheet_filename = f'{FOLDER_BASE}/{customer}/{project}/timesheet_teste_{project}.xlsx'

    wb = load_workbook(timesheet_filename)
    try:
        ws = wb['total_hours']
    except KeyError:
        ws = wb.create_sheet('total_hours')

    new_row = 2

    labels = (
        'data',
        'mês',
        'tempo',
        'tempo_display',
        'issues',
        'sprint',
    )

    bold_calibri = styles.Font(bold=True, name='Calibri')

    # Set labels and apply font in a loop
    for col, label in enumerate(labels, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = bold_calibri

    for item in total_hours:
        ws.cell(row=new_row, column=1, value=item['date'])

        cell = ws.cell(row=new_row, column=2, value=item['month'])
        cell.alignment = styles.Alignment(horizontal='center')

        ws.cell(row=new_row, column=3, value=item['total_hours'])
        ws.cell(row=new_row, column=4, value=item['total_hours_display'])
        ws.cell(row=new_row, column=5, value=item['issues'])

        cell = ws.cell(row=new_row, column=6, value=item['sprint'])
        cell.alignment = styles.Alignment(horizontal='center')

        new_row += 1

    wb.save(timesheet_filename)


def export_timesheet_service(project):
    tasks = project.get_tasks()

    customer = project.customer.name
    project = project.title
    timesheet_filename = f'{FOLDER_BASE}/{customer}/{project}/timesheet_teste_{project}.xlsx'

    try:
        wb = load_workbook(timesheet_filename)
        ws = wb['timesheet']
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.create_sheet('timesheet')
        wb.save(timesheet_filename)

    new_row = 2

    labels = (
        'data',
        'hora_inicial',
        'hora_final',
        'tempo',
        'tempo_display',
        'issue',
        'titulo',
    )

    bold_calibri = styles.Font(bold=True, name='Calibri')

    # Set labels and apply font in a loop
    for col, label in enumerate(labels, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = bold_calibri

    for task in tasks:
        for timesheet in task.get_timesheets():
            ws.cell(row=new_row, column=1, value=timesheet.date_from_start_time_display)
            ws.cell(row=new_row, column=2, value=timesheet.start_time_display)
            ws.cell(row=new_row, column=3, value=timesheet.end_time_display)
            ws.cell(row=new_row, column=4, value=timesheet.get_hour())
            ws.cell(row=new_row, column=5, value=timesheet.get_hour_display())

            cell = ws.cell(row=new_row, column=6, value=timesheet.task.issue.number)
            cell.alignment = styles.Alignment(horizontal='center')

            ws.cell(row=new_row, column=7, value=timesheet.task.issue.title)

            new_row += 1

    wb.save(timesheet_filename)

    total_hours = group_by_date(project)

    write_total_hours_on_timesheet_file(customer, project, total_hours)